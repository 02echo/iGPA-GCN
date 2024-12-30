import json
import openpyxl

with open('geneId_num_map.json', "r") as f:
    gene_map = json.load(f)
    print("获取gene编号字典成功....")

wb = openpyxl.load_workbook('geneIdnetwork.xlsx')
ws = wb['Sheet1']
for i in range(895642):
    cur_gene1 = ws.cell(i+1,1).value
    cur_gene2 = ws.cell(i+1,3).value
    gene1_number = gene_map[str(cur_gene1)]
    gene2_number = gene_map[str(cur_gene2)]
    ws.cell(i+1,2).value = gene1_number
    ws.cell(i+1,4).value = gene2_number
    if i % 10000 == 0:
        print("当前处理第{}行完成".format(i))
wb.save('geneIdnetwork.xlsx')
wb.close()
print('编号写入成功！')