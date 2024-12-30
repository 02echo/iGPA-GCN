
import pandas as pd
import xlrd
import numpy as np
import openpyxl

df = pd.read_excel('gene.xlsx')
gene_list = df['gene_symbol'].values.tolist()
gene_list = list(set(gene_list))
# print(len(gene_list))

workbook = xlrd.open_workbook('gene.xlsx')
sheet = workbook.sheets()[0]

# 写入数据
wb = openpyxl.load_workbook('gene.xlsx')
ws = wb.get_sheet_by_name('Sheet1')

# for i in range(1,20397):
#     cur_gene_list = sheet.cell_value(i,0).split(' ')
#     for cur in cur_gene_list:
#         if cur in gene_list:
#             c = ws.cell(i+1,4)
#             c.value = '存在'
#             break
#     if i%1000 == 0:
#         print("当前处理完第{}行".format(i))
# wb.save('gene.xlsx')


all_raw_gene = []
for i in range(1,20397):
    cur_row = sheet.cell_value(i,0).split(' ')
    all_raw_gene.append(cur_row)
all_gene = [x for j in all_raw_gene for x in j]
all_gene = list(set(all_gene))

for i in range(1,1255):
    cur_gene = sheet.cell_value(i,1)
    if cur_gene in all_gene:
        c = ws.cell(i+1,5)
        c.value = 'yes'
    if i%100 == 0:
        print("当前处理完第{}行".format(i))
wb.save('gene.xlsx')

