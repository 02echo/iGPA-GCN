
import pandas as pd
import numpy as np
import json
import openpyxl


def get_gene_list():
    with open('my_unique_gene_list.json', 'r') as f:
        load_list = json.load(f)
    print("获取gene列表成功。")
    return load_list


def process_gene_list(j, is_first):
    wb = openpyxl.load_workbook('gene2gene.xlsx')
    print('start...')
    ws = wb['Sheet1']
    visited = []
    pre = ""
    for i in range(1, 895643):
        cur_row = ws.cell(i+1,j+1).value.split(' ')
        if len(cur_row) == 1:
            continue
        elif not is_first:
            flag = False
            for gene in cur_row:
                if gene in gene_list:
                    ws.cell(i+1, j+1).value = gene
                    flag = True
                    break
            if not flag:
                ws.cell(i+1, j+1).value = cur_row[0]
        elif is_first and cur_row not in visited:
            flag = False
            visited.append(cur_row)
            for gene in cur_row:
                if gene in gene_list:
                    pre = gene
                    flag = True
                    break
            c = ws.cell(i + 1, j + 1)
            if not flag:
                pre = cur_row[0]
            c.value = pre
        else:
            c = ws.cell(i + 1, j + 1)
            c.value = pre
        if i % 1000 == 0:
            print("当前处理完第{}行".format(i))
    wb.save('gene2gene.xlsx')
    wb.close()


gene_list = get_gene_list()
# 处理第一列gene,第一列gene相同的在一起
# process_gene_list(0, True)
# 处理第二列gene，第二列gene相同的不在一起，不需要visited数组
process_gene_list(1, False)


