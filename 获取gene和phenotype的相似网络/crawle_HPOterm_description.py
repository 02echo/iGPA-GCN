import requests
# from bs4 import BeautifulSoup
import json
import openpyxl


wb = openpyxl.load_workbook('phenotype_information.xlsx')
ws = wb['Sheet1']
# 添加headers
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0',
           'Connection': 'close'}

description = []
synonyms = []
parents = []
# for i in range(2,5848):
# for i in range(2,1500):
# for i in range(1500,2500):
# for i in range(2500,3072):
# for i in range(3072,4000):
for i in range(4000,5848):
# for i in range(4500, 5848):
    cur_phe = ws.cell(i,2).value
    cur_url = 'https://ontology.jax.org/api/hp/terms/' + cur_phe
    # 以下是描述与同义词的爬取
    # try:
    #     r = requests.get(cur_url, headers=headers, timeout=10)
    # except:
    #     for j in range(4):
    #         res = requests.get(cur_url, headers=headers, timeout=20)
    #         if res.status_code == 200:
    #             break
    # cur_json = json.loads(r.text)
    # description.append(cur_json['definition'])
    # synonyms.append(str(cur_json['synonyms']) if cur_json['synonyms'] else '无')

    try:
        # parents = requests.get(cur_url + '/parents', headers=headers, timeout=10)
        children = requests.get(cur_url + '/children', headers=headers, timeout=10)
    except:
        for j in range(4):
            # res = requests.get(cur_url + '/parents', headers=headers, timeout=20)
            res = requests.get(cur_url + '/children', headers=headers, timeout=20)
            if res.status_code == 200:
                break

    ws.cell(i,7).value = children.text
    wb.save('phenotype_information.xlsx')
    wb.close()
    if i % 10 == 0:
        print('目前处理完第{}条数据'.format(i))

# with open('phenotype_parents1.json','w') as f:
#     json.dump(description, f)
#     print('表型父节点数据保存成功！')

# with open('phenotype_synonyms8.json', 'w') as f:
#     json.dump(synonyms, f)
#     print('表型同义词数据保存成功！')
# print(len(description))
# print(len(synonyms))


# 读取保存的json文件
def load_json(json_file):
    with open(json_file, 'r') as f:
        data = json.load(f)
        print('文件读取成功!')
    return data


# ws.cell(i,4).value = description[i-2]
# ws.cell(i,5).value = synonyms[i-2]
# ws.cell(i,6).value = parents.text
# ws.cell(i,7).value = children.text

# 写入表格数据
def write_xlsx(data, index):
    # 读取表型信息
    wb = openpyxl.load_workbook('phenotype_information.xlsx')
    ws = wb['Sheet1']
    for i in range(2,5848):
        ws.cell(i,index).value = data[i-2]
    print('数据写入完成！')
    wb.save('phenotype_information.xlsx')
    wb.close()


# data = load_json('merged_des.json')
# index = 4
# data = load_json('merged_syn.json')
# index = 5
# write_xlsx(data, index)









