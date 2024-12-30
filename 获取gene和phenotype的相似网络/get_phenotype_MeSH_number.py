# coding=utf-8
import re
import sys
import json
import openpyxl
import ast
import numpy as np
from ast import literal_eval

wb = openpyxl.load_workbook('phenotype_information.xlsx')
ws = wb['Sheet1']

with open('mesh_word.json', 'r') as f:
    mesh_list = json.load(f)
    print(len(mesh_list))
    print('数据载入成功！')


# 保存出现在表型HPO中的所有Mesh项
mesh_dic = set()



# 处理字符串中的标点
def remove_chars(str):
    chars = '[·’!"\#$%&\'()＃！（）*+,-./:;<=>?\@，：?￥★、…．＞【】［］《》？“”‘’\[\\]^_`{|}~]+'
    res = re.sub(chars, "", str)
    return res


# 处理所有的列表
def trans2Big(c):
    str_list = []
    for ch in c.split(' '):
        ch = ch.capitalize()
        str_list.append(ch)
    cur_word = ' '.join(str_list)
    return cur_word


def update_weight_matrix(matrix, row_index, mesh_term):
    col_index = mesh_arr.index(mesh_term)
    cur_weight = matrix[row_index][col_index]
    cur_weight += 1
    matrix[row_index][col_index] = cur_weight

# cur = ws.cell(14,1).value
# if cur in mesh_arr:
#     print('hhh00')
# if cur in mesh_list:
#     print('hh01')
# # cur_syn = 'Abnormality of the musculature of the lower limbs'.split(' ')
# # for c in cur_syn:
# #     if c in mesh_arr:
# #         print('hhh111')
# #     if c in mesh_list:
# #         print('hhhh222')
# cur_description = ws.cell(14, 4).value
# if cur_description:
#     cur_description = ws.cell(14,4).value.split(' ')
#     # num = 0
#     for c in cur_description:
#         if c in mesh_arr:
#             print('hhh222')
#         if c in mesh_list:
#             print('hhh00222')
# cur_list = ws.cell(14,6).value
# if cur_list != '[]':
#     cur_list = json.loads(cur_list)
#
#     # num = 0
#     for k in range(len(cur_list)):
#         cur_word = cur_list[k]['name']
#         if cur_word in mesh_arr:
#             print('hhh333')
#         if cur_word in mesh_list:
#             print('hhh000333')
# cur_list = ws.cell(14,7).value
# if cur_list != '[]':
#     cur_list = cur_list.split('"id":')
#     for k in range(1, len(cur_list)):
#         cur_word = cur_list[k].split('translations')[0][21:-3]
#         if cur_word in mesh_arr:
#             print('hhhh444')
#         if cur_word in mesh_list:
#             print('hhh000444')


def get_hpo_name():
    #处理本词
    for i in range(2, 5848):
    # for i in range(2, 2104):
        cur_hpo = ws.cell(i, 1).value
        if cur_hpo in mesh_list:
            # cur_num = ws.cell(i, 8).value
            # ws.cell(i, 8).value = cur_num + 1
            if cur_hpo not in mesh_dic:
                mesh_dic.add(cur_hpo)
            update_weight_matrix(mesh_weight_matrix, i-2, cur_hpo)

        if i % 100 == 0:
            print('当前处理完第{}行数据'.format(i))
    print('#########################')
    print('处理完本词')


# 处理同义词
def get_hpo_syn():
    for i in range(2, 5848):
        cur_syn = ws.cell(i, 5).value
        more_num = 0
        if cur_syn != '无':
            cur_syn = literal_eval(cur_syn)
            for c in cur_syn:
                cur_word = trans2Big(c)
                if cur_word in mesh_list:
                    more_num += 1
                    if cur_word not in mesh_dic:
                        mesh_dic.add(cur_word)
                    update_weight_matrix(mesh_weight_matrix, i-2, cur_word)
            # cur_num = ws.cell(i, 8).value
            # ws.cell(i, 8).value = cur_num + more_num

        if i % 100 == 0:
            print('当前处理完第{}行数据'.format(i))
    print('#########################')
    print('处理完同义词')


# 处理hpo描述
def get_hpo_description():
    for i in range(2, 5848):
        cur_description = ws.cell(i, 4).value
        num = 0
        if cur_description:
            cur_description = remove_chars(ws.cell(i, 4).value)
            cur_description = cur_description.split(' ')
            for c in cur_description:
                c = c.capitalize()
                if c in mesh_list:
                    num += 1
                    update_weight_matrix(mesh_weight_matrix, i-2, c)
                    if c not in mesh_dic:
                        mesh_dic.add(c)

        # cur_num = ws.cell(i, 8).value
        # ws.cell(i, 8).value = cur_num + num

        if i % 100 == 0:
            print('当前处理完第{}行数据'.format(i))
    print('#########################')
    print('处理完hpo描述')


def get_hpo_parent():
    for i in range(2, 5848):
        cur_list = ws.cell(i, 6).value
        if cur_list != '[]':
            cur_list = json.loads(cur_list)

            num = 0
            for k in range(len(cur_list)):
                cur_word_list = cur_list[k]['name'].split(' ')
                for c in cur_word_list:
                    cur_word = trans2Big(c)
                    if cur_word in mesh_list:
                        num += 1
                        update_weight_matrix(mesh_weight_matrix, i-2, cur_word)
                        if cur_word not in mesh_dic:
                            mesh_dic.add(cur_word)

            # cur_num = ws.cell(i, 8).value
            # ws.cell(i, 8).value = cur_num + num

        if i % 100 == 0:
            print('当前处理完第{}行数据'.format(i))
    print('#########################')
    print('处理完parent')


def get_hpo_children():
    for i in range(2, 5848):
        cur_list = ws.cell(i,7).value
        if cur_list != '[]':
            cur_list = cur_list.split('"id":')

            num = 0
            for k in range(1, len(cur_list)):
                cur_word = cur_list[k].split('translations')[0][21:-3]
                cur_word = trans2Big(cur_word)
                if cur_word in mesh_list:
                    num += 1
                    update_weight_matrix(mesh_weight_matrix, i-2, cur_word)
                    if cur_word not in mesh_dic:
                        mesh_dic.add(cur_word)

            # cur_num = ws.cell(i, 8).value
            # ws.cell(i, 8).value = cur_num + num

        if i % 100 == 0:
            print('当前处理完第{}行数据'.format(i))
    print('#########################')
    print('处理完children')


# def get_Mesh_weight_matrix():


# get_hpo_name()
# get_hpo_syn()
# get_hpo_description()
# get_hpo_parent()
# get_hpo_children()
#
#
# mesh_arr = list(mesh_dic)
# with open('occurred_mesh.json', 'w') as f:
#     json.dump(mesh_arr, f)
#     print('数据保存成功！')
#     print(len(mesh_arr))
#
# # parent 6,  children 7
# wb.save('phenotype_information.xlsx')
# wb.close()

with open('occurred_mesh.json', 'r') as f:
    mesh_arr = json.load(f)
    print(len(mesh_arr))


mesh_weight_matrix = np.zeros((5846, 3703))

get_hpo_name()
print(mesh_weight_matrix[39].__contains__(1))
get_hpo_syn()
print(mesh_weight_matrix[39].__contains__(1))
get_hpo_description()
print(mesh_weight_matrix[39].__contains__(1))
get_hpo_parent()
print(mesh_weight_matrix[39].__contains__(1))
get_hpo_children()
print(mesh_weight_matrix[39].__contains__(1))

# print(test39)

np.save('mesh_weight_matrix.npy', mesh_weight_matrix)
print('保存mesh项权重矩阵成功！')



