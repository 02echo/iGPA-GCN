import json
import openpyxl
import numpy as np


# 读取gene-id信息
with open('gene_id.json', 'r') as f:
    gene_map = json.load(f)
    print('基因数据载入成功！')

print(len(gene_map))

sim_matrix = np.eye(1254)
wb = openpyxl.load_workbook('ppi.xlsx')
ws = wb['Sheet1']

gene_not_list = ['ADPRHL2', 'C11orf80', 'CCDC84', 'CTCFL', 'DDX58', 'FAM126A', 'FAM92A', 'KIAA0556', 'MAATS1', 'MAP11', 'PIH1D3', 'TTC26', 'WDR34', 'WDR60', 'WDR66']


for i in range(2,6883):
    node1 = ws.cell(i,1).value
    node2 = ws.cell(i,2).value
    if node1 not in gene_not_list and node2 not in gene_not_list:
        node1_index = gene_map[node1]
        node2_index = gene_map[node2]
        sim_matrix[node1_index][node2_index] = ws.cell(i,3).value

np.save('gene_simNetwork.npy', sim_matrix)
print('基因相似矩阵保存成功！')

