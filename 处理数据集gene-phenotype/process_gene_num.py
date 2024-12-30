import openpyxl
import numpy as np
import pandas as pd
import json

wb = openpyxl.load_workbook('gene_information.xlsx')
ws = wb['Sheet1']
gene_map = {}
pre = 1254
for i in range(2,1256):
    cur_id = ws.cell(i,2).value
    cur_num = ws.cell(i,3).value
    gene_map[str(cur_id)] = cur_num


for i in range(16606):
    cur_id = ws.cell(i+1,5).value
    if str(cur_id) not in gene_map:
        gene_map[str(cur_id)] = pre
        pre += 1
    ws.cell(i+1,6).value = gene_map[str(cur_id)]

wb.save('gene_information.xlsx')
wb.close()

with open('geneID_num_map.json','w') as f:
    json.dump(gene_map,f)
    print("载入文件完成！")
print(len(gene_map))
