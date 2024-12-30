import numpy as np
import openpyxl
import json


# 获取gene和phenotype的map信息
with open('GP_gene_num_map.json') as f:
    gene_num_map = json.load(f)
print('geneMap文件载入成功！')

with open('GP_phenotype_num_map.json') as f:
    phenotype_num_map = json.load(f)
print('phenotypeMap文件载入成功！')

wb = openpyxl.load_workbook('gene_phenotype_network.xlsx')
ws = wb['Sheet1']

# 填入gene和phenotype对应的num
for i in range(2,69051):
    cur_gene_id = ws.cell(i,1).value
    cur_phe_id = ws.cell(i,3).value
    ws.cell(i,2).value = gene_num_map[str(cur_gene_id)]
    ws.cell(i,4).value = phenotype_num_map[str(cur_phe_id)]

wb.save('gene_phenotype_network.xlsx')
wb.close()
