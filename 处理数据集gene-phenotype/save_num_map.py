import json
import openpyxl


# PD_gene_num_map.json
wb = openpyxl.load_workbook('gene_information.xlsx')
ws = wb['Sheet1']
gene_num_map = {}

for i in range(2,1256):
    cur_id = ws.cell(i, 2).value
    cur_num = ws.cell(i, 3).value
    gene_num_map[str(cur_id)] = cur_num

with open('GP_gene_num_map.json','w') as f:
    json.dump(gene_num_map, f)
    print('文件载入成功！')
print(len(gene_num_map))


# PD_phenotype_num_map.json
# wb = openpyxl.load_workbook('phenotype_information.xlsx')
# ws = wb['Sheet1']
# phenotype_num_map = {}
#
# for i in range(2, 5848):
#     cur_id = ws.cell(i, 2).value
#     cur_num = ws.cell(i, 3).value
#     phenotype_num_map[str(cur_id)] = cur_num
#
#
# print(len(phenotype_num_map))
# with open('GP_phenotype_num_map.json','w') as f:
#     json.dump(phenotype_num_map, f)
#     print('文件载入成功！')
