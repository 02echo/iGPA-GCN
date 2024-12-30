import json
import numpy as np
import openpyxl


result = np.load('predicted_score.npy')
score = []
gene_list = []
phe_list = []

for i in range(300):
    for j in range(30):
        if result[i][j] > 1.8:
            score.append(str(result[i][j]))
            phe_list.append(i)
            gene_list.append(j)


web_res = []
wb = openpyxl.load_workbook('gene_information.xlsx')
ws = wb['Sheet1']

wb1 = openpyxl.load_workbook('phenotype_information.xlsx')
ws1 = wb1['Sheet1']

for i in range(17):
    cur_gene = ws.cell(gene_list[i]+2,1).value
    cur_phenotype = ws1.cell(phe_list[i]+2,1).value
    print(cur_phenotype)
    cur_score = score[i]
    web_res.append({'gene': cur_gene, 'phenotype': cur_phenotype, 'score': cur_score})

# for i in range(len(web_res)):
#     # print(web_res[i],',', "\n")
#     print(cur)





